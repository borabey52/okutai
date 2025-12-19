import streamlit as st
import utils
import pandas as pd
import io
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.chart import BarChart, Reference

# --- SAYFA VE MERKEZİ YÖNETİM ---
st.set_page_config(page_title="Analiz", page_icon="📊", layout="wide")
utils.sayfa_yukle() 
# --------------------------------

st.title("📊 Sınav Arşivi ve Analiz")

# Yenileme Butonu
col_ref1, col_ref2 = st.columns([0.8, 0.2])
with col_ref2:
    if st.button("🔄 Arşivi Güncelle", use_container_width=True):
        st.session_state.sinif_verileri = utils.load_results(st.session_state.user_id)
        st.toast("Veriler güncellendi!", icon="✅")

if not st.session_state.sinif_verileri:
    st.info("Henüz veri yok.")
    st.stop()

# --- SINAV SEÇİMİ VE SİLME İŞLEMİ ---
df_tum = pd.DataFrame(st.session_state.sinif_verileri)
benzersiz_oturumlar = df_tum['Oturum'].unique()

st.markdown("### 🔍 Sınav Seçimi")

# Alanı ikiye bölüyoruz: Seçim Kutusu (%85) - Silme Butonu (%15)
col_sel, col_del = st.columns([0.85, 0.15], gap="small")

with col_sel:
    secilen_oturum = st.selectbox("İncelemek istediğiniz sınavı seçin:", benzersiz_oturumlar, label_visibility="collapsed")

with col_del:
    # Popover: Tıklayınca küçük bir pencere açar (Yanlışlıkla silmeyi engeller)
    with st.popover("🗑️ Sil", use_container_width=True):
        st.warning(f"'{secilen_oturum}' sınavı kalıcı olarak silinecek!")
        if st.button("Evet, Sil", type="primary", use_container_width=True):
            if utils.delete_exam(st.session_state.user_id, secilen_oturum):
                st.toast(f"{secilen_oturum} silindi.", icon="🗑️")
                # Session state'i güncelle ve sayfayı yenile
                st.session_state.sinif_verileri = utils.load_results(st.session_state.user_id)
                st.rerun()
            else:
                st.error("Silinemedi.")

# Filtreleme
df = df_tum[df_tum['Oturum'] == secilen_oturum]
filtrelenmis_liste = df.to_dict('records')
df_tablo = utils.tablo_olustur(filtrelenmis_liste)

st.divider()

# --- 1. GENEL BAKIŞ ---
st.header("1. Genel Bakış")
c1, c2, c3, c4 = st.columns(4)
c1.metric("Öğrenci Sayısı", len(df))
c2.metric("Sınıf Ortalaması", f"{df['Toplam Puan'].mean():.1f}")
c3.metric("En Yüksek Not", f"{int(df['Toplam Puan'].max())}")
basari = int(len(df[df['Toplam Puan'] >= 50])/len(df)*100) if len(df) > 0 else 0
c4.metric("Başarı Oranı", f"%{basari}")
st.markdown("---")

# --- 2. NOT ÇİZELGESİ ---
st.header("2. Öğrenci Not Çizelgesi")
st.dataframe(df_tablo, use_container_width=True)
st.markdown("---")

# --- 3. YANITLAR VE YORUMLAR ---
st.header("3. Yanıtlar ve Yorumlar")
for ogr in filtrelenmis_liste:
    p = int(ogr.get('Toplam Puan', 0))
    ik = "🌟" if p >= 85 else "✅" if p >= 50 else "⚠️"
    
    with st.expander(f"{ik} {ogr.get('Ad Soyad', '?')} | {p} Puan"):
        if "Detaylar" in ogr:
            for soru in ogr["Detaylar"]:
                p_val = float(soru.get('puan', 0))
                t_val = float(soru.get('tam_puan', 0))
                
                renk_kod = "green" if p_val == t_val and t_val > 0 else "red" if p_val == 0 else "orange"
                ikon = "✅" if p_val == t_val and t_val > 0 else "❌" if p_val == 0 else "⚠️"
                
                p_text = f"{int(p_val)}" if p_val == int(p_val) else f"{p_val}"
                t_text = f"{int(t_val)}" if t_val == int(t_val) else f"{t_val}"

                st.markdown(f"""
                <div style="font-size:18px; margin-bottom:5px;">
                    <strong>Soru {soru.get('no')}</strong> {ikon} <span style="color:{renk_kod}; font-weight:bold;">[{p_text} / {t_text}]</span>
                </div>
                <div style="font-size:16px; margin-bottom:10px; color:#333;">
                    <strong>Cevap:</strong> {soru.get('cevap')}
                </div>
                <div style="background-color:#f0f8ff; padding:15px; border-radius:8px; border-left:6px solid #002D62; font-size:16px;">
                    <span style="font-weight:bold; color:#002D62;">🤖 Yorum:</span> {soru.get('yorum')}
                </div>
                <hr style="margin: 10px 0;">
                """, unsafe_allow_html=True)
st.markdown("---")

# --- 4. GRAFİKLER ---
st.header("4. Grafikler")
c_g1, c_g2 = st.columns(2)
with c_g1:
    st.markdown("**Soru Başarı Analizi**")
    scols = [c for c in df_tablo.columns if "Soru " in c]
    if scols: st.bar_chart(df_tablo[scols].apply(pd.to_numeric).mean(), color="#4CAF50")
    else: st.warning("Detay yok.")
with c_g2:
    st.markdown("**Not Dağılımı**")
    df['Durum'] = pd.cut(df['Toplam Puan'], [0,44,69,84,100], labels=['Zayıf','Orta','İyi','Pekiyi'], include_lowest=True)
    st.bar_chart(df['Durum'].value_counts(), color="#FF9800")
st.markdown("---")

# --- EXCEL OLUŞTURMA ---
st.header("📥 Rapor İndir")

detay_listesi = []
merge_bilgisi = []
satir_sayaci = 2 

for ogr in filtrelenmis_liste:
    sorular = ogr.get("Detaylar", [])
    if not sorular: continue
    baslangic_satiri = satir_sayaci
    for soru in sorular:
        p_val = float(soru.get('puan', 0))
        t_val = float(soru.get('tam_puan', 0))
        yorum_metni = soru.get("yorum")
        if p_val == t_val and t_val > 0: yorum_metni = "" 
        
        detay_listesi.append({
            "Öğrenci Adı": ogr.get("Ad Soyad"),
            "Numara": ogr.get("Numara"),
            "Genel Puan": int(ogr.get("Toplam Puan", 0)),
            "Soru": soru.get("no"),
            "Puan": soru.get("puan"),
            "OkutAİ Yorumu": yorum_metni
        })
        satir_sayaci += 1
    
    if (satir_sayaci - 1) > baslangic_satiri:
        merge_bilgisi.append((baslangic_satiri, satir_sayaci - 1))

df_detay = pd.DataFrame(detay_listesi)

scols = [c for c in df_tablo.columns if "Soru " in c]
soru_basarisi = df_tablo[scols].apply(pd.to_numeric).mean().reset_index()
soru_basarisi.columns = ['Soru', 'Ortalama']
dagilim = df['Durum'].value_counts().sort_index().reset_index()
dagilim.columns = ['Aralık', 'Öğrenci Sayısı']

buffer = io.BytesIO()
with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
    df_tablo.to_excel(writer, index=False, sheet_name='Not Çizelgesi')
    df_detay.to_excel(writer, index=False, sheet_name='Detaylı Yorumlar')
    soru_basarisi.to_excel(writer, index=False, sheet_name='Grafik_Veri', startrow=0, startcol=0)
    dagilim.to_excel(writer, index=False, sheet_name='Grafik_Veri', startrow=0, startcol=4)
    
    ws1 = writer.sheets['Not Çizelgesi']
    ws_data = writer.sheets['Grafik_Veri']
    
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws1.iter_rows(min_row=1, max_row=len(df_tablo)+1):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                if cell.column == 2:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws1.column_dimensions['B'].width = 25
    
    chart1 = BarChart(); chart1.type = "col"; chart1.style = 10; chart1.title = "Soru Başarı Analizi"; chart1.varyColors = True; chart1.legend = None 
    data1 = Reference(ws_data, min_col=2, min_row=1, max_row=len(soru_basarisi)+1)
    cats1 = Reference(ws_data, min_col=1, min_row=2, max_row=len(soru_basarisi)+1)
    chart1.add_data(data1, titles_from_data=True); chart1.set_categories(cats1) 
    ws1.add_chart(chart1, f"A{len(df_tablo)+4}")
    
    chart2 = BarChart(); chart2.type = "col"; chart2.style = 10; chart2.title = "Puan Dağılımı"; chart2.varyColors = True; chart2.legend = None
    data2 = Reference(ws_data, min_col=5, min_row=1, max_row=len(dagilim)+1)
    cats2 = Reference(ws_data, min_col=4, min_row=2, max_row=len(dagilim)+1)
    chart2.add_data(data2, titles_from_data=True); chart2.set_categories(cats2) 
    ws1.add_chart(chart2, f"F{len(df_tablo)+4}")

    ws2 = writer.sheets['Detaylı Yorumlar']
    for bas, bit in merge_bilgisi:
        ws2.merge_cells(start_row=bas, start_column=1, end_row=bit, end_column=1)
        ws2.merge_cells(start_row=bas, start_column=2, end_row=bit, end_column=2)
        ws2.merge_cells(start_row=bas, start_column=3, end_row=bit, end_column=3)
        
    ws2.column_dimensions['A'].width = 25; ws2.column_dimensions['F'].width = 75
    
    for row in ws2.iter_rows(min_row=1):
        for cell in row:
            if cell.row == 1:
                cell.font = header_font; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
            else:
                cell.border = thin_border
                if cell.column == 1 or cell.column == 6: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else: cell.alignment = Alignment(horizontal='center', vertical='center')

st.download_button(
    label=f"📥 {secilen_oturum} Raporunu İndir",
    data=buffer.getvalue(),
    file_name=f'{secilen_oturum}_Raporu.xlsx',
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    type="primary",
    use_container_width=True
)
st.caption("ℹ️ Not: Excel dosyasını indirmeden önce, açıksa eski dosyayı kapatınız.")
# ... (Üst kısımlar aynı)
# Kodun en sonuna şunu ekle:
utils.footer_ekle()